from datetime import datetime

import xlrd
from openpyxl import Workbook, load_workbook

'''
#创建写入文件
'''
# wb = Workbook()
# ws = wb.active
# ws['A1'] = 42
# ws.append([1,2,3])
# ws['A2'] = datetime.now()
# wb.save('data/sample.xlsx')
'''
xlrd读取文件
'''
# book = xlrd.open_workbook('data/data.xlsx')
# sheet = book.sheet_by_index(0)
# print('第一行内容：',sheet.row_values(0))
# print('第二行内容：',sheet.row_values(1))
# print('第三行内容：',sheet.row_values(2))

'''
openpyxl读取文件
'''
wb = load_workbook(filename='data/data.xlsx')
ws = wb['topics']
# print(ws['B3'].value) #读取单个单元格
test_data = []
for x in range(2,4):
    testcase_data = []
    for y in range(2,7):
        testcase_data.append(ws.cell(row=x,column=y).value)
        print(ws.cell(row=x,column=y).value)
    test_data.append(testcase_data)
print(test_data)






