import requests
from openpyxl import load_workbook

"""
常用功能：
1.发布话题
"""
accesstoken = 'ebb04940-7910-4f74-a175-fe5dd01cb0f9'
other_token = '81ccd1d9-be88-43b6-b288-b17293300f32'
defaulturl = 'http://49.233.108.117:3000/api/v1'

def Create_topic(topic_data):
    """
    获取topic_id
    :param topic_data:
    :return:
    """
    url1 = defaulturl+'/topics'
    res = requests.post(url=url1, data=topic_data).json()
    topic_id = res['topic_id']
    return topic_id

def Get_topic_detail(update_topicid):
    """
    获取话题详情
    :param update_topicid:
    :return:
    """
    url = defaulturl+'/topic/'+update_topicid
    res = requests.get(url)
    return res

def Get_username(topic_data):
    """
    用户发帖-topic_id查看话题详情-获取loginname
    :param topic_data:
    :return:
    """
    url1 = defaulturl + '/topics'
    res = requests.post(url=url1, data=topic_data).json()
    topic_id = res['topic_id']
    url = defaulturl+'/topic/'+topic_id
    res1 = requests.get(url).json()
    username = res1['data']['author']['loginname']
    return username

def Get_reply_id(topic_id,replydata):
    """
    获取评论ID
    :param topic_id:
    :param replydata:
    :return:
    """
    url = defaulturl+'/topic/'+topic_id+'/replies'
    res = requests.post(url=url,data=replydata).json()
    return res['reply_id']

def Get_verfily_token(url,data):
    """
    获取验证accesstoken的键值
    :param url:
    :param data:
    :return:
    """
    # url = defaulturl+'accesstoken'
    res = requests.post(url=url,data=data).json()
    success = res['success']
    loginname = res['loginname']
    id = res['id']
    return success,loginname,id

def Get_hasnot_number(url,data):
    """
    统计未读消息计数
    :param url:
    :param data:
    :return:
    """
    res = requests.get(url=url,data=data).json()
    success = res['success']
    data1 = res['data']
    return success,data1

def Get_allmessageNum(url,data):
    """
    获取已读|未读消息各多少
    :param url:
    :param data:
    :return:
    """
    res = requests.get(url=url, data=data).json()
    read = len(res['data']['has_read_messages'])
    notread = len(res['data']['hasnot_read_messages'])
    return read, notread

def Get_markPartMessage(data):
    """
    获取未读消息的第一条
    :param data:
    :return:
    """
    url = defaulturl+'/messages'
    res = requests.get(url=url,data=data).json()
    part_message = res['data']['hasnot_read_messages'][0]['id']
    return part_message
def Get_token():
    """
    获取最新的token值
    :return:
    """
    return accesstoken

def Other_token():
    """
    获取他人token，用来评论
    :return:
    """
    return other_token
def GetExceldata(filename,sheetname):
    '''
    openpyxl读取文件
    '''
    wb = load_workbook(filename)
    ws = wb[sheetname]
    # print(ws['B3'].value) #读取单个单元格
    print(len(tuple(ws.rows)))
    test_data = []
    for x in range(2, len(tuple(ws.rows)) + 1):
        testcase_data = []
        for y in range(2, 7):
            testcase_data.append(ws.cell(row=x, column=y).value)
            print(ws.cell(row=x, column=y).value)
        test_data.append(testcase_data)
    return test_data